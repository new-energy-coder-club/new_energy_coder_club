#!/usr/bin/env python3
"""读取「工商业光伏电站分层分级_两步走」xlsx 模板，输出结构化字典。

主要用途：
1. 把模板的示例区（简版 R39-R44 / 复杂版 R47-R53）转成 score.py 可消费的项目数据。
2. 让 Claude 在用户给出 xlsx 路径时快速看清模板里有哪些项目行。

用法：
    python parse_template.py "D:\\OneDrive\\Desktop\\副本工商业光伏电站分层分级_两步走.xlsx"
"""
from __future__ import annotations
import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


SHEET1_NAME_HINT = "第一步"
SHEET2_NAME_HINT = "第二步"

SHEET1_COLUMNS = ["项目名称", "维度1", "维度2", "维度3", "维度4", "维度5", "维度6",
                  "综合评级", "基准IRR", "加减项", "最终IRR"]
SHEET2_COLUMNS = ["项目名称", "维度1", "维度2", "维度3", "维度4",
                  "维度5a", "维度5b", "维度5c", "维度5d", "维度6",
                  "基础总分", "加减分", "最终总分", "等级", "基准IRR"]


def _find_sheet(wb, hint):
    for n in wb.sheetnames:
        if hint in n:
            return wb[n]
    raise KeyError(f"找不到含「{hint}」的 sheet，现有：{wb.sheetnames}")


def _find_example_block(ws, header_keyword):
    """找出包含 header_keyword 的标题行，下面所有有「项目名称」+维度的行都视作示例。"""
    header_row = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and header_keyword in str(v):
                header_row = r
                break
        if header_row:
            break
    if header_row is None:
        return None, []
    name_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(header_row + 1, c).value == "项目名称":
            name_col = c
            break
    if name_col is None:
        return None, []
    rows = []
    r = header_row + 2
    while r <= ws.max_row:
        name = ws.cell(r, name_col).value
        if not name:
            break
        row_vals = [ws.cell(r, c).value for c in range(name_col, ws.max_column + 1)]
        rows.append(row_vals)
        r += 1
    return header_row, rows


_GRADE_RE = re.compile(r"^([ABCD])(?:\s*/\s*([\d.]+))?$")


def _parse_grade_cell(v):
    if v is None:
        return None, None
    s = str(v).strip()
    m = _GRADE_RE.match(s)
    if m:
        return m.group(1), float(m.group(2)) if m.group(2) else None
    if s in {"A", "B", "C", "D"}:
        return s, None
    return None, None


def parse_workbook(fp: str | Path):
    wb = openpyxl.load_workbook(str(fp), data_only=False)
    out = {"sheets": {}, "simple_examples": [], "complex_examples": []}

    ws1 = _find_sheet(wb, SHEET1_NAME_HINT)
    out["sheets"]["simple"] = ws1.title
    _, rows1 = _find_example_block(ws1, "简版项目评估示例")
    for row in rows1:
        if len(row) < 7:
            continue
        proj = {
            "name": row[0],
            "grades": {
                "d1_enterprise": _parse_grade_cell(row[1])[0],
                "d2_region":     _parse_grade_cell(row[2])[0],
                "d3_self_use":   _parse_grade_cell(row[3])[0],
                "d4_materials":  _parse_grade_cell(row[4])[0],
                "d5_finance":    _parse_grade_cell(row[5])[0],
                "d6_industry":   _parse_grade_cell(row[6])[0],
            },
            "raw_excel": {
                "综合评级":    row[7] if len(row) > 7 else None,
                "基准IRR":     row[8] if len(row) > 8 else None,
                "加减项":      row[9] if len(row) > 9 else None,
                "最终IRR":     row[10] if len(row) > 10 else None,
            },
        }
        out["simple_examples"].append(proj)

    ws2 = _find_sheet(wb, SHEET2_NAME_HINT)
    out["sheets"]["complex"] = ws2.title
    _, rows2 = _find_example_block(ws2, "复杂版项目评估示例")
    for row in rows2:
        if len(row) < 10:
            continue
        proj = {
            "name": row[0],
            "grades": {
                "d1_enterprise":       _parse_grade_cell(row[1])[0],
                "d2_region":           _parse_grade_cell(row[2])[0],
                "d3_self_use":         _parse_grade_cell(row[3])[0],
                "d4_materials":        _parse_grade_cell(row[4])[0],
                "d5a_finance_health":  _parse_grade_cell(row[5])[0],
                "d5b_pv_share":        _parse_grade_cell(row[6])[0],
                "d5c_revenue_cover":   _parse_grade_cell(row[7])[0],
                "d5d_profit_cover":    _parse_grade_cell(row[8])[0],
                "d6_industry":         _parse_grade_cell(row[9])[0],
            },
            "raw_excel": {
                "基础总分":    row[10] if len(row) > 10 else None,
                "加减分":      row[11] if len(row) > 11 else None,
                "最终总分":    row[12] if len(row) > 12 else None,
                "等级":        row[13] if len(row) > 13 else None,
                "基准IRR":     row[14] if len(row) > 14 else None,
            },
        }
        out["complex_examples"].append(proj)

    return out


def main():
    p = argparse.ArgumentParser(description="读取两步走 xlsx 模板")
    p.add_argument("path", help="xlsx 文件路径")
    p.add_argument("--out", help="保存为 JSON")
    args = p.parse_args()

    data = parse_workbook(args.path)
    s = json.dumps(data, ensure_ascii=False, indent=2)
    print(s)
    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(s)


if __name__ == "__main__":
    main()
