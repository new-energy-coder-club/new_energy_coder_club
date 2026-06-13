#!/usr/bin/env python3
"""把评分结果写回「两步走」xlsx 模板。

默认行为：
- 在原 xlsx 旁边生成 `<原文件>_filled.xlsx`，**不会修改原件**。
- 在两个示例区（简版 R39 之后 / 复杂版 R47 之后）各追加一行评分结果。
- 保留原表的合并单元格、列宽、表头格式。

用法：
    python fill_template.py "副本工商业光伏电站分层分级_两步走.xlsx" \
        --data project.json \
        --out filled.xlsx          # 可选，默认 <原>_filled.xlsx
    python fill_template.py template.xlsx --data a.json --data b.json --in-place

如果想直接覆盖原文件，加 --in-place（会先备份为 .bak.xlsx）。
"""
from __future__ import annotations
import argparse
import json
import shutil
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font
except ImportError:
    print("需要 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).parent))
from score import score_project, DIM_WEIGHTS, D5_SUB  # noqa: E402


SIMPLE_HEADER_KEY = "简版项目评估示例"
COMPLEX_HEADER_KEY = "复杂版项目评估示例"
FILLED_FILL = PatternFill(start_color="FFF7E6", end_color="FFF7E6", fill_type="solid")
FILLED_FONT = Font(bold=False, color="0050A0")


def _find_block_end(ws, header_keyword: str, name_col: int = 1):
    """返回（标题行号, 最后一行有数据的行号）。"""
    header_row = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and header_keyword in str(v):
                header_row = r
                break
        if header_row:
            break
    if header_row is None:
        raise KeyError(f"找不到含「{header_keyword}」的标题行")
    last_data_row = header_row + 1
    r = header_row + 2
    while r <= ws.max_row and ws.cell(r, name_col).value:
        last_data_row = r
        r += 1
    return header_row, last_data_row


def _fmt_irr(lo: float, hi: float) -> str:
    return f"{lo:g}%-{hi:g}%"


def _fmt_adj_simple(result: dict) -> str:
    items = result["adjustments"]
    if not items:
        return "无"
    parts = []
    for a in items:
        sign = "+" if a["irr_delta_pct"] >= 0 else ""
        parts.append(f"{a['name']}（{sign}{a['irr_delta_pct']}%）")
    return "；".join(parts)


def _fmt_adj_complex(result: dict) -> str:
    delta = result["step2_weighted"]["adjustment_total"]
    return f"{delta:+g}" if delta else "0"


def _style_row(ws, row_no: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row_no, c)
        cell.fill = FILLED_FILL
        cell.font = FILLED_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def write_simple_row(ws, result: dict):
    header, last = _find_block_end(ws, SIMPLE_HEADER_KEY)
    new_row = last + 1
    g = result["step1_simple"]["per_dim_grades"]
    method = result["step1_simple"]["method_used"].split("（")[0]
    grade = result["final_grade"]
    base_lo = result["irr"]["baseline_low"]
    base_hi = result["irr"]["baseline_high"]
    final_lo = result["irr"]["final_low"]
    final_hi = result["irr"]["final_high"]

    vals = [
        result["name"],
        g["d1_enterprise"],
        g["d2_region"],
        g["d3_self_use"],
        g["d4_materials"],
        g["d5_finance"],
        g["d6_industry"],
        f"{grade}级（{method}）",
        _fmt_irr(base_lo, base_hi),
        _fmt_adj_simple(result),
        _fmt_irr(final_lo, final_hi),
    ]
    for i, v in enumerate(vals, start=1):
        ws.cell(new_row, i).value = v
    _style_row(ws, new_row, len(vals))
    return new_row


def write_complex_row(ws, result: dict):
    header, last = _find_block_end(ws, COMPLEX_HEADER_KEY)
    new_row = last + 1

    s2 = result["step2_weighted"]
    sub = s2["d5_sub_scores"]
    dim = s2["dim_scores"]
    g = result["grades_used"]

    def cell(grade_key, score):
        return f"{g[grade_key]}/{score:g}"

    base = s2["base_total"]
    adj = s2["adjustment_total"]
    final = s2["final_total"]
    grade = result["final_grade"]
    irr_lo = result["irr"]["final_low"]
    irr_hi = result["irr"]["final_high"]

    vals = [
        result["name"],
        cell("d1_enterprise", dim["d1_enterprise"]),
        cell("d2_region",     dim["d2_region"]),
        cell("d3_self_use",   dim["d3_self_use"]),
        cell("d4_materials",  dim["d4_materials"]),
        cell("d5a_finance_health", sub["d5a_finance_health"]),
        cell("d5b_pv_share",       sub["d5b_pv_share"]),
        cell("d5c_revenue_cover",  sub["d5c_revenue_cover"]),
        cell("d5d_profit_cover",   sub["d5d_profit_cover"]),
        cell("d6_industry",   dim["d6_industry"]),
        f"{base:g}",
        f"{adj:+g}" if adj else "0",
        f"{final:g}",
        f"{grade}级",
        _fmt_irr(irr_lo, irr_hi),
    ]
    for i, v in enumerate(vals, start=1):
        ws.cell(new_row, i).value = v
    _style_row(ws, new_row, len(vals))
    return new_row


def fill_workbook(template_path: str | Path, projects: list[dict],
                  out_path: str | Path | None = None, in_place: bool = False) -> Path:
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(template_path)

    if in_place:
        backup = template_path.with_suffix(".bak.xlsx")
        shutil.copy2(template_path, backup)
        target = template_path
        print(f"[i] 已备份原文件 → {backup}")
    else:
        if out_path is None:
            target = template_path.with_name(template_path.stem + "_filled.xlsx")
        else:
            target = Path(out_path)
        shutil.copy2(template_path, target)

    wb = openpyxl.load_workbook(str(target))
    ws_simple = None
    ws_complex = None
    for n in wb.sheetnames:
        if "第一步" in n:
            ws_simple = wb[n]
        elif "第二步" in n:
            ws_complex = wb[n]
    if ws_simple is None or ws_complex is None:
        raise KeyError(f"模板缺少「第一步/第二步」sheet：{wb.sheetnames}")

    summary = []
    for data in projects:
        r = score_project(data)
        s_row = write_simple_row(ws_simple, r)
        c_row = write_complex_row(ws_complex, r)
        summary.append({
            "name": r["name"],
            "simple_row": s_row,
            "complex_row": c_row,
            "final_grade": r["final_grade"],
            "final_total": r["step2_weighted"]["final_total"],
            "irr": f"{r['irr']['final_low']:g}%-{r['irr']['final_high']:g}%",
        })

    wb.save(str(target))
    return target, summary


def main():
    p = argparse.ArgumentParser(description="把评分结果写回 xlsx 模板示例区")
    p.add_argument("template", help="两步走 xlsx 模板路径")
    p.add_argument("--data", action="append", required=True,
                   help="项目数据 JSON 文件（可重复多次以批量评分）")
    p.add_argument("--out", help="输出 xlsx 路径，默认 <原>_filled.xlsx")
    p.add_argument("--in-place", action="store_true", help="覆盖原文件（自动备份为 .bak.xlsx）")
    args = p.parse_args()

    projects = []
    for fp in args.data:
        with open(fp, "r", encoding="utf-8") as f:
            projects.append(json.load(f))

    target, summary = fill_workbook(args.template, projects, args.out, args.in_place)

    print(f"\n[✓] 已写入：{target}\n")
    print(f"{'项目':<30s}  {'简版行':>6s}  {'复杂行':>6s}  {'等级':<5s}  {'总分':>6s}  {'IRR'}")
    print("-" * 80)
    for s in summary:
        print(f"{s['name'][:28]:<30s}  {s['simple_row']:>6d}  {s['complex_row']:>6d}  "
              f"{s['final_grade']:<5s}  {s['final_total']:>6g}  {s['irr']}")


if __name__ == "__main__":
    main()
